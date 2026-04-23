import cv2
import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
correct_predictions = 0
total_predictions = 0 
 
THRESHOLD = 85
# =========================
# Create folders
# =========================
os.makedirs("captured_images", exist_ok=True)

# =========================
# Load Face Detector
# =========================
face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")

# =========================
# Load Recognizer
# =========================
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainer.yml")

# Load ID + Names from dataset
names = {}
for folder in os.listdir("dataset"):
    try:
        sid = int(folder.split("_")[0])
        name = folder.split("_")[1]
        names[sid] = name
    except:
        continue

# =========================
# Excel Setup
# =========================
excel_file = "attendance.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Name", "Date", "Time", "Photo"])
    wb.save(excel_file)

wb = load_workbook(excel_file)
ws = wb.active

# =========================
# Open Camera
# =========================
cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)

if not cam.isOpened():
    print("Camera not detected")
    exit()

print("Camera opened... Detecting face")

captured = False

while True:
    ret, frame = cam.read()
    if not ret:
        continue

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        total_predictions+=1
        face_gray = gray[y:y+h, x:x+w]
        face_gray = cv2.resize(face_gray, (200, 200))
        face_gray = cv2.equalizeHist(face_gray)

        sid, conf = recognizer.predict(face_gray)
        total_predictions+=1 
        
        if conf < THRESHOLD:
            person_name = names.get(sid, "Unknown")
            correct_predictions+= 1
        else:
            person_name = "Unknown"

        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
        cv2.putText(frame, person_name, (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        # ===== Capture only once and prevent duplicate =====
        if person_name != "Unknown" and not captured:
            now = datetime.now()
            date = now.strftime("%Y-%m-%d")
            time = now.strftime("%H-%M-%S")

            # Check duplicate (same ID + same date)
            duplicate = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == sid and row[2] == date:
                    duplicate = True
                    break

            if duplicate:
                print("Already marked today — Duplicate prevented")
                captured = True
                break

            captured = True

            # Write on image
            cv2.putText(frame, f"{sid} {person_name} {date} {time}",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX,
                        0.8, (0, 0, 255), 2)

            # Save image
            img_path = f"captured_images/{sid}_{person_name}_{date}_{time}.jpg"
            cv2.imwrite(img_path, frame)
            print("Image captured")

            # ===== Save to Excel =====
            ws.append([sid, person_name, date, time, ""])

            # Insert image into Excel
            img_excel = XLImage(img_path)
            img_excel.width = 80
            img_excel.height = 60
            ws.add_image(img_excel, f"E{ws.max_row}")
           
            wb.save(excel_file)
            print("Attendance saved in Excel")

    cv2.imshow("Auto Attendance", frame)

    if captured:
        cv2.waitKey(2000)
        break

    if cv2.waitKey(1) == 27:
        break

cam.release()
cv2.destroyAllWindows()

if total_predictions > 0:
    accuracy = (correct_predictions / total_predictions) * 100
    
    print("\n------ PERFORMANCE REPORT ------")
    print("Total Faces Processed:", total_predictions)
    print("Successful Recognitions:", correct_predictions)
    print("Accuracy: {:.2f}%".format(accuracy))

    # Count number of registered students
    num_students = len(names)

    # Save to CSV for graph
    with open("accuracy_data.csv", "a") as f:
        f.write(f"{num_students},{accuracy}\n")
print("Process Completed")
